"""Defines the `xlsx` class."""

from typing import Optional
from printly.style import get_rgb_values
from printly.types import Color
from ..table import Table


class xlsx:  # pylint: disable=invalid-name
    """Represents XLSX io operations."""

    @staticmethod
    def load(table: Table, filepath: str) -> None:  # pylint: disable=too-many-locals
        """Imports table rows from excel file."""
        from openpyxl import load_workbook  # type: ignore  # pylint: disable=import-outside-toplevel
        from ..properties import (  # pylint: disable=import-outside-toplevel
            Border,
            Background,
            Font,
            Text,
        )

        def convert_color(color) -> Optional[Color]:
            if color:
                return "#" + color.rgb[2:]
            return None

        def convert_border(xborder) -> Border:
            styles = {
                "hair": "thin",
                "thin": "single",
                "medium": "solid",
                "thick": "solid",
                "dashed": "dashed",
                "dotted": "dotted",
                "double": "double",
            }
            border = Border(None, None)
            for side in ("left", "right", "top", "bottom"):
                getattr(border, side).style = styles.get(getattr(xborder, side).border_style)
                getattr(border, side).color = convert_color(getattr(xborder, side).color)
            return border

        def convert_font(font, hidden: bool) -> Font:
            font_style = ""
            if font.bold:
                font_style += "bold+"
            if font.italic:
                font_style += "italic+"
            if font.strike:
                font_style += "strikethrough+"
            if font.underline == "single":
                font_style += "underline+"
            elif font.underline == "double":
                font_style += "double-underline+"
            if hidden:
                font_style += "hidden+"
            return Font(font_style.rstrip("+") or None, convert_color(font.color))

        def convert_text(value, alignment) -> Text:
            return Text(
                text=value,
                justify=(
                    "left" if alignment.horizontal == "general" else alignment.horizontal or "left"
                ),
                align=alignment.vertical or "top",
                wrap=alignment.wrap_text,
            )

        workbook = load_workbook(filepath)
        worksheet = workbook.active
        for x in range(1, worksheet.max_row + 1):
            cells = []
            for y in range(1, worksheet.max_column + 1):
                xcell = worksheet.cell(row=x, column=y)
                cells.append(
                    {
                        "value": xcell.value,
                        "background": Background(convert_color(xcell.fill.bgColor)),
                        "border": convert_border(xcell.border),
                        "font": convert_font(xcell.font, xcell.protection.hidden),
                        "text": convert_text(xcell.value, xcell.alignment),
                    }
                )
            table.add_row((cell["value"] for cell in cells))
            for index, cell in enumerate(cells):
                table[-1][index].background = cell["background"]
                table[-1][index].border = cell["border"]
                table[-1][index].font = cell["font"]
                table[-1][index].text = cell["text"]

    @staticmethod
    def dump(table: Table, filepath: str) -> None:  # pylint: disable=too-many-locals
        """Exports a table to excel file."""
        from openpyxl import Workbook  # type: ignore  # pylint: disable=import-outside-toplevel
        from openpyxl.styles import (  # type: ignore  # pylint: disable=import-outside-toplevel
            Alignment,
            Border,
            Font,
            PatternFill,
            Protection,
            Side,
        )

        def convert_color(color: Color) -> str:
            return "ff" + "".join((f"{int(v):02x}" for v in get_rgb_values(color)))

        def convert_border_side(border_side) -> Side:
            style = {
                None: None,
                "single": "thin",
                "double": "double",
                "dashed": "dashed",
                "dotted": "dotted",
                "solid": "medium",
            }.get(border_side.style)
            color = convert_color(border_side.color) if border_side.color else "ff000000"
            return Side(style, color)

        workbook = Workbook()
        worksheet = workbook.active
        alphabet_letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

        for y, column in enumerate(table._columns):  # pylint: disable=protected-access
            for x, cell in enumerate(column, start=1):
                xcell = worksheet[alphabet_letters[y] + str(x)]
                xcell.value = cell.value
                bgcolor = (
                    convert_color(cell.background.color) if cell.background.color else "ffffffff"
                )
                fgcolor = convert_color(cell.font.color) if cell.font.color else "ff000000"
                xcell.fill = PatternFill(start_color=bgcolor, end_color=bgcolor, fill_type="solid")
                xcell.font = Font(size=12, color=fgcolor)
                if cell.font.style:
                    xcell.font.bold = "bold" in cell.font.style
                    xcell.font.italic = "italic" in cell.font.style
                    xcell.font.strike = "strikethrough" in cell.font.style
                    xcell.font.underline = (
                        "double"
                        if "double-underline" in cell.font.style
                        else "single" if "underline" in cell.font.style else "none"
                    )
                    if "hidden" in cell.font.style:
                        xcell.protection = Protection(hidden=True)
                xcell.border = Border(
                    left=convert_border_side(cell.border.left),
                    right=convert_border_side(cell.border.right),
                    top=convert_border_side(cell.border.top),
                    bottom=convert_border_side(cell.border.bottom),
                )
                xcell.alignment = Alignment(
                    horizontal=cell.text.justify, vertical=cell.text.align, wrap_text=cell.text.wrap
                )
            max_width = max((cell.width for cell in column))
            max_padding = max((sum(cell.padding.inline) for cell in column))
            worksheet.column_dimensions[alphabet_letters[y]].width = max_width + max_padding + 1
        workbook.save(filepath)
